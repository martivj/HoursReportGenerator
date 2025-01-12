import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from pandas import DataFrame

import subprocess

from typing import List, Dict, Tuple
from src.types.dataclasses import GroupSummary, ProjectInfo


class ExcelFormatter:

    def __init__(
        self,
        wb: Workbook,
        data: List[
            Tuple[
                ProjectInfo,
                Dict[str, DataFrame],  # Group data
                Dict[str, GroupSummary],  # Group summaries
            ]
        ],
        output_path: str,
        total_sheet_first: bool = True,
        close_open_excel: bool = True,
        style_vars: Dict[str, int] = None,
    ):
        self.wb = wb
        self.data = data
        self.output_path = output_path
        self.total_sheet_first = total_sheet_first
        self.close_open_excel = close_open_excel
        self.style_vars = style_vars or self._get_default_style_vars()

    def _get_default_style_vars(self) -> Dict[str, int]:
        return {
            # Group sheet dimensions
            "part_col_width": 25,
            "week_col_width": 10,
            "date_col_width": 15,
            "minutes_col_width": 10,
            "description_col_width": 100,
            "summary_metric_col_width": 15,
            "summary_value_col_width": 15,
            "parts_row_height": 32,
            # Total sheet dimensions
            "project_col_width": 25,
            "total_hours_col_width": 15,
            "total_weeks_col_width": 15,
            "avg_hours_per_week_col_width": 25,
            "tab_color": "FFD700",
            "total_row_height": 32,
        }

    def _close_excel(self, gracefully: bool = True) -> bool:
        """Close the running Excel session if it's open."""
        # Skip if not on Windows
        if not self.close_open_excel:
            return False

        try:
            import win32com.client as win32
        except ImportError:
            print("win32com not available - Excel closing disabled")
            return False

        print("")
        if gracefully:
            try:
                excel = win32.GetActiveObject("Excel.Application")
                excel.Quit()
                print("Closed running Excel session gracefully.")
                return True
            except Exception as e:
                print(f"Failed to close Excel gracefully: {e}")
                return False
        else:
            try:
                result = subprocess.call(["taskkill", "/F", "/IM", "EXCEL.EXE"])
                if result == 0:  # Success
                    print("Closed running Excel session forcefully.")
                    return True
                else:
                    print("Excel was not running")
                    return False
            except Exception as e:
                print(f"Failed to close Excel: {e}")
                return False

    def format(self) -> None:
        """Format and save the Excel workbook with all project data."""
        # Close any open Excel instances first
        excel_was_running = False
        if self.close_open_excel:
            excel_was_running = self._close_excel(gracefully=True)

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        if self.total_sheet_first:
            self._format_total_sheet()
            self._format_part_sheets()
        else:
            self._format_part_sheets()
            self._format_total_sheet()

        # Save and close workbook
        print("Saving the workbook...")
        self.wb.save(self.output_path)
        self.wb.close()

        # Reopen the workbook
        if self.close_open_excel:
            if excel_was_running:
                print("Reopening Excel file...")
            else:
                print("Opening Excel file...")
            os.startfile(self.output_path)

    def _format_part_sheets(self) -> None:
        """Format individual sheets for each project's parts."""
        for project_info, parts_data, summaries in self.data:
            for part_name, df in parts_data.items():
                # Create sheet
                sheet_name = f"{project_info.title} {part_name}"
                ws = self.wb.create_sheet(title=sheet_name)

                # Set tab color
                ws.sheet_properties.tabColor = project_info.primary_color

                # Write column headers first (row 2, after title row)
                for col, header in enumerate(df.columns, start=1):
                    ws.cell(row=2, column=col, value=header)

                # Write data starting from row 3
                for idx, row in df.iterrows():
                    for col, value in enumerate(row, start=1):
                        ws.cell(
                            row=idx + 3,
                            column=col,
                            value=value,  # +3 for title and header rows
                        )

                # Format sheet
                self._format_part_sheet(
                    ws=ws,
                    title=f"{project_info.title} - {part_name}",
                    primary_color=project_info.primary_color,
                    secondary_color=project_info.secondary_color,
                    total_hours=summaries[part_name].total_hours,
                    hours_per_week=summaries[part_name].hours_per_week,
                )

    def _format_part_sheet(
        self,
        ws,
        title: str,
        primary_color: str,
        secondary_color: str,
        total_hours: float,
        hours_per_week: DataFrame,
    ) -> None:
        """Format a single part sheet with data and styling."""
        # Add title
        ws.merge_cells("A1:E1")
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style header row
        self._format_header_row(ws, 2, primary_color)

        # Format data section
        self._format_data_section(ws, secondary_color)

        # Add and format weekly overview
        self._add_weekly_overview(ws, hours_per_week, primary_color, secondary_color)

        # Add and format summary section
        self._add_summary_section(
            ws, total_hours, hours_per_week, primary_color, secondary_color
        )

    def _format_total_sheet(self) -> None:
        """Create and format the total summary sheet."""
        ws = self.wb.create_sheet(title="Total")
        ws.sheet_properties.tabColor = self.style_vars["tab_color"]
        ws.sheet_view.showGridLines = False

        col_offset = 0
        row = 1

        for i, (project_info, _, summaries) in enumerate(self.data):
            # Calculate position in 2x2 grid
            if i % 3 == 0 and i > 0:  # Every 2nd dataset starts a new row
                row += 12  # Height of each project section
                col_offset = 0

            # Merge cells for title (A-D or F-I depending on column offset)
            start_col = chr(ord("A") + col_offset)
            end_col = chr(ord("D") + col_offset)
            ws.merge_cells(f"{start_col}{row}:{end_col}{row}")

            # Add project section title
            title_cell = ws.cell(
                row=row, column=1 + col_offset, value=f"{project_info.title} - Summary"
            )
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set row height for title
            ws.row_dimensions[row].height = self.style_vars["total_row_height"]

            # Add part summaries and totals with column offset
            next_row = self._add_project_summary(
                ws, row + 1, project_info, summaries, col_offset
            )
            next_row = self._add_project_totals_summary(
                ws, next_row + 1, project_info, summaries, col_offset
            )

            # Move to next column position
            col_offset += 5  # Leave space between columns

    def _format_header_row(self, ws, row: int, color: str) -> None:
        """Format the header row with given color."""
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _format_data_section(self, ws, alternate_color: str) -> None:
        """Format the main data section with alternating colors."""
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        # Add border to all cells
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply alternating colors and center alignment
        light_fill = PatternFill(
            start_color=alternate_color, end_color=alternate_color, fill_type="solid"
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in range(3, max_row + 1):  # Start after header
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align

                # Apply alternating row colors
                if row % 2 == 0:
                    cell.fill = light_fill

        # Set column widths
        ws.column_dimensions["A"].width = self.style_vars["part_col_width"]
        ws.column_dimensions["B"].width = self.style_vars["week_col_width"]
        ws.column_dimensions["C"].width = self.style_vars["date_col_width"]
        ws.column_dimensions["D"].width = self.style_vars["minutes_col_width"]
        ws.column_dimensions["E"].width = self.style_vars["description_col_width"]

        # Apply row height to all rows
        for row in range(1, max_row + 1):
            ws.row_dimensions[row].height = self.style_vars["parts_row_height"]

    def _add_weekly_overview(
        self, ws, hours_per_week: DataFrame, primary_color: str, secondary_color: str
    ) -> None:
        """Add and format weekly hours overview section."""
        # Start position for weekly overview
        start_col = 7  # Column G
        start_row = 2

        # Add headers
        ws.cell(row=start_row, column=start_col, value="Week")
        ws.cell(row=start_row, column=start_col + 1, value="Hours")

        # Style headers
        header_fill = PatternFill(
            start_color=primary_color, end_color=primary_color, fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col in range(start_col, start_col + 2):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        light_fill = PatternFill(
            start_color=secondary_color, end_color=secondary_color, fill_type="solid"
        )

        for idx, row in hours_per_week.iterrows():
            current_row = start_row + idx + 1

            # Week number
            week_cell = ws.cell(
                row=current_row, column=start_col, value=f"Week {row['week']}"
            )

            # Hours value
            hours_cell = ws.cell(
                row=current_row, column=start_col + 1, value=row["duration_hours"]
            )
            hours_cell.number_format = "0.0"

            # Format both cells
            for cell in [week_cell, hours_cell]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if current_row % 2 == 0:
                    cell.fill = light_fill

        # Set column widths
        ws.column_dimensions[get_column_letter(start_col)].width = self.style_vars[
            "summary_metric_col_width"
        ]
        ws.column_dimensions[get_column_letter(start_col + 1)].width = self.style_vars[
            "summary_value_col_width"
        ]

        # Apply row height to all rows in weekly overview
        for row in range(start_row, start_row + len(hours_per_week) + 1):
            ws.row_dimensions[row].height = self.style_vars["parts_row_height"]

    def _add_summary_section(
        self,
        ws,
        total_hours: float,
        hours_per_week: DataFrame,
        primary_color: str,
        secondary_color: str,
    ) -> None:
        """Add and format summary section with totals."""
        # Start position (below weekly overview)
        start_col = 7  # Column G
        start_row = len(hours_per_week) + 4  # Add some spacing

        # Add header
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=start_col + 1,
        )
        header = ws.cell(row=start_row, column=start_col, value="Summary")
        header.fill = PatternFill(
            start_color=primary_color, end_color=primary_color, fill_type="solid"
        )
        header.font = Font(bold=True, color="FFFFFF")
        header.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Add summary data
        data = [
            ("Total Hours", total_hours),
            ("Total Weeks", len(hours_per_week)),
            (
                "Average Hours/Week",
                total_hours / len(hours_per_week) if len(hours_per_week) > 0 else 0,
            ),
        ]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        light_fill = PatternFill(
            start_color=secondary_color, end_color=secondary_color, fill_type="solid"
        )

        for idx, (metric, value) in enumerate(data, 1):
            current_row = start_row + idx

            # Metric name
            metric_cell = ws.cell(row=current_row, column=start_col, value=metric)

            # Value
            value_cell = ws.cell(row=current_row, column=start_col + 1, value=value)
            value_cell.number_format = "0.0" if "Hours" in metric else "0"

            # Format both cells
            for cell in [metric_cell, value_cell]:
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                if current_row % 2 == 0:
                    cell.fill = light_fill

        # Set row height
        for row in range(start_row, start_row + len(data) + 1):
            ws.row_dimensions[row].height = self.style_vars["parts_row_height"]

    def _add_project_summary(
        self,
        ws,
        start_row: int,
        project_info: ProjectInfo,
        summaries: Dict[str, GroupSummary],
        col_offset: int = 0,
    ) -> int:
        # Add headers with offset
        headers = ["Project", "Total Hours", "Total Weeks", "Average Hours/Week"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col + col_offset, value=header)
            cell.fill = PatternFill(
                start_color=project_info.primary_color,
                end_color=project_info.primary_color,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data for each part
        current_row = start_row + 1
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for part_name, summary in summaries.items():
            row_data = [
                part_name,
                summary.total_hours,
                len(summary.hours_per_week),
                (
                    summary.total_hours / len(summary.hours_per_week)
                    if len(summary.hours_per_week) > 0
                    else 0
                ),
            ]

            for col, value in enumerate(row_data, 1):
                # Add col_offset here
                cell = ws.cell(row=current_row, column=col + col_offset, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                if current_row % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=project_info.secondary_color,
                        end_color=project_info.secondary_color,
                        fill_type="solid",
                    )

                if col in [2, 4]:  # Hours columns
                    cell.number_format = "0.0"
                elif col == 3:  # Weeks column
                    cell.number_format = "0"

            current_row += 1

        # Set column widths with offset
        ws.column_dimensions[get_column_letter(1 + col_offset)].width = self.style_vars[
            "project_col_width"
        ]
        ws.column_dimensions[get_column_letter(2 + col_offset)].width = self.style_vars[
            "total_hours_col_width"
        ]
        ws.column_dimensions[get_column_letter(3 + col_offset)].width = self.style_vars[
            "total_weeks_col_width"
        ]
        ws.column_dimensions[get_column_letter(4 + col_offset)].width = self.style_vars[
            "avg_hours_per_week_col_width"
        ]

        # Apply row height to all rows in project summary
        for row in range(start_row, current_row):
            ws.row_dimensions[row].height = self.style_vars["total_row_height"]

        return current_row

    def _add_project_totals_summary(
        self,
        ws,
        start_row: int,
        project_info: ProjectInfo,
        summaries: Dict[str, GroupSummary],
        col_offset: int = 0,
    ) -> int:
        """Add summary totals for entire project. Returns next row number."""
        # Calculate project totals
        total_hours = sum(s.total_hours for s in summaries.values())
        total_weeks = sum(len(s.hours_per_week) for s in summaries.values())
        avg_hours = total_hours / total_weeks if total_weeks > 0 else 0

        # Add header with offset
        ws.merge_cells(
            start_row=start_row,
            start_column=1 + col_offset,
            end_row=start_row,
            end_column=2 + col_offset,
        )
        header = ws.cell(row=start_row, column=1 + col_offset, value="Project Total")
        header.fill = PatternFill(
            start_color=project_info.primary_color,
            end_color=project_info.primary_color,
            fill_type="solid",
        )
        header.font = Font(bold=True, color="FFFFFF")
        header.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Add summary data with offset
        data = [
            ("Total Hours", total_hours),
            ("Total Weeks", total_weeks),
            ("Average Hours/Week", avg_hours),
        ]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for idx, (metric, value) in enumerate(data, 1):
            current_row = start_row + idx

            # Metric name with offset
            metric_cell = ws.cell(row=current_row, column=1 + col_offset, value=metric)

            # Value with offset
            value_cell = ws.cell(row=current_row, column=2 + col_offset, value=value)
            value_cell.number_format = "0.0" if "Hours" in metric else "0"

            # Format both cells
            for cell in [metric_cell, value_cell]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if current_row % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=project_info.secondary_color,
                        end_color=project_info.secondary_color,
                        fill_type="solid",
                    )

        # Apply row heights
        for row in range(start_row, start_row + len(data) + 1):
            ws.row_dimensions[row].height = self.style_vars["total_row_height"]

        return start_row + len(data) + 1
